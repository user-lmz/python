#!/usr/bin/env python

import openpyxl

book = openpyxl.load_workbook('sheets.xlsx')

book.create_sheet("April")

print(book.sheetnames)

sheet1 = book["January"]
book.remove(sheet1)

print(book.sheetnames)

book.create_sheet("January", 0)
book.create_sheet("May", 3)
print(book.sheetnames)

book.save('sheets2.xlsx')